
import csv
import shutil
import sqlite3
from datetime import datetime, date
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

DB_NAME = "inventario_salud_v43.db"
DATE_FMT = "%Y-%m-%d"
DATETIME_FMT = "%Y-%m-%d %H:%M:%S"
DEFAULT_USER = "admin"
DEFAULT_PASS = "admin123"
BACKUP_DIR = "respaldos"

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except Exception:
    OPENPYXL_OK = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False


class Database:
    def __init__(self, db_path=DB_NAME):
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path)
        self.conn.row_factory = sqlite3.Row
        self.create_tables()
        self.seed_default_user()

    def now(self):
        return datetime.now().strftime(DATETIME_FMT)

    def create_tables(self):
        cur = self.conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS usuarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password TEXT NOT NULL,
                nombre_completo TEXT NOT NULL,
                rol TEXT NOT NULL DEFAULT 'Administrador',
                activo INTEGER NOT NULL DEFAULT 1,
                creado_en TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS insumos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT UNIQUE,
                nombre TEXT NOT NULL,
                categoria TEXT NOT NULL,
                stock_actual INTEGER NOT NULL DEFAULT 0,
                stock_minimo INTEGER NOT NULL DEFAULT 0,
                unidad TEXT NOT NULL,
                vencimiento TEXT,
                ubicacion TEXT,
                proveedor TEXT,
                lote TEXT,
                observaciones TEXT,
                creado_en TEXT NOT NULL,
                actualizado_en TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS movimientos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                insumo_id INTEGER NOT NULL,
                tipo TEXT NOT NULL,
                cantidad INTEGER NOT NULL,
                motivo TEXT,
                usuario TEXT NOT NULL,
                fecha TEXT NOT NULL,
                stock_resultante INTEGER NOT NULL,
                FOREIGN KEY (insumo_id) REFERENCES insumos(id)
            )
        """)
        self.conn.commit()

    def seed_default_user(self):
        cur = self.conn.cursor()
        cur.execute("SELECT id FROM usuarios WHERE username=?", (DEFAULT_USER,))
        if cur.fetchone() is None:
            cur.execute("""
                INSERT INTO usuarios (username, password, nombre_completo, rol, creado_en)
                VALUES (?, ?, ?, ?, ?)
            """, (DEFAULT_USER, DEFAULT_PASS, "Administrador General", "Administrador", self.now()))
            self.conn.commit()

    def authenticate_user(self, username, password):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT * FROM usuarios
            WHERE username=? AND password=? AND activo=1
        """, (username.strip(), password.strip()))
        return cur.fetchone()

    def change_password(self, username, current_password, new_password):
        user = self.authenticate_user(username, current_password)
        if user is None:
            raise ValueError("La contraseña actual no es correcta.")
        if len(new_password.strip()) < 4:
            raise ValueError("La nueva contraseña debe tener al menos 4 caracteres.")
        cur = self.conn.cursor()
        cur.execute("UPDATE usuarios SET password=? WHERE username=?", (new_password.strip(), username))
        self.conn.commit()

    def create_backup(self, backup_folder=BACKUP_DIR):
        backup_path = Path(backup_folder)
        backup_path.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        destination = backup_path / f"respaldo_inventario_v43_{timestamp}.db"
        self.conn.commit()
        self.conn.close()
        shutil.copy2(self.db_path, destination)
        self.conn = sqlite3.connect(self.db_path)
        self.conn.row_factory = sqlite3.Row
        return str(destination)

    def get_item_by_id(self, item_id):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM insumos WHERE id=?", (item_id,))
        return cur.fetchone()

    def get_item_by_code(self, codigo):
        cur = self.conn.cursor()
        cur.execute("SELECT * FROM insumos WHERE codigo=?", (codigo,))
        return cur.fetchone()

    def add_item(self, data, usuario):
        now = self.now()
        cur = self.conn.cursor()
        cur.execute("""
            INSERT INTO insumos (
                codigo, nombre, categoria, stock_actual, stock_minimo, unidad,
                vencimiento, ubicacion, proveedor, lote, observaciones,
                creado_en, actualizado_en
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            data["codigo"], data["nombre"], data["categoria"], data["stock_actual"],
            data["stock_minimo"], data["unidad"], data["vencimiento"], data["ubicacion"],
            data["proveedor"], data["lote"], data["observaciones"], now, now
        ))
        item_id = cur.lastrowid
        self.conn.commit()
        if data["stock_actual"] > 0:
            self.add_movement(item_id, "Entrada", data["stock_actual"], "Stock inicial del registro", usuario)
        return item_id

    def update_item(self, item_id, data):
        cur = self.conn.cursor()
        cur.execute("""
            UPDATE insumos
            SET codigo=?, nombre=?, categoria=?, stock_minimo=?, unidad=?, vencimiento=?,
                ubicacion=?, proveedor=?, lote=?, observaciones=?, actualizado_en=?
            WHERE id=?
        """, (
            data["codigo"], data["nombre"], data["categoria"], data["stock_minimo"],
            data["unidad"], data["vencimiento"], data["ubicacion"], data["proveedor"],
            data["lote"], data["observaciones"], self.now(), item_id
        ))
        self.conn.commit()

    def set_stock_direct(self, item_id, new_stock):
        cur = self.conn.cursor()
        cur.execute("UPDATE insumos SET stock_actual=?, actualizado_en=? WHERE id=?", (new_stock, self.now(), item_id))
        self.conn.commit()

    def delete_item(self, item_id):
        cur = self.conn.cursor()
        cur.execute("DELETE FROM movimientos WHERE insumo_id=?", (item_id,))
        cur.execute("DELETE FROM insumos WHERE id=?", (item_id,))
        self.conn.commit()

    def fetch_items(self, search_text="", category="Todas"):
        cur = self.conn.cursor()
        sql = "SELECT * FROM insumos"
        clauses = []
        params = []
        if search_text.strip():
            like = f"%{search_text.strip()}%"
            clauses.append("(codigo LIKE ? OR nombre LIKE ? OR categoria LIKE ? OR ubicacion LIKE ? OR proveedor LIKE ? OR lote LIKE ?)")
            params.extend([like, like, like, like, like, like])
        if category != "Todas":
            clauses.append("categoria=?")
            params.append(category)
        if clauses:
            sql += " WHERE " + " AND ".join(clauses)
        sql += " ORDER BY nombre COLLATE NOCASE ASC"
        cur.execute(sql, params)
        return cur.fetchall()

    def get_categories(self):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT DISTINCT categoria FROM insumos
            WHERE categoria IS NOT NULL AND categoria <> ''
            ORDER BY categoria
        """)
        return [row[0] for row in cur.fetchall()]

    def add_movement(self, item_id, tipo, cantidad, motivo, usuario):
        item = self.get_item_by_id(item_id)
        if item is None:
            raise ValueError("El insumo seleccionado no existe.")
        cantidad = int(cantidad)
        if cantidad <= 0:
            raise ValueError("La cantidad debe ser mayor que 0.")
        stock_actual = int(item["stock_actual"])
        if tipo == "Salida":
            if cantidad > stock_actual:
                raise ValueError("No hay stock suficiente para registrar esta salida.")
            nuevo_stock = stock_actual - cantidad
        else:
            nuevo_stock = stock_actual + cantidad
        cur = self.conn.cursor()
        cur.execute("UPDATE insumos SET stock_actual=?, actualizado_en=? WHERE id=?", (nuevo_stock, self.now(), item_id))
        cur.execute("""
            INSERT INTO movimientos (insumo_id, tipo, cantidad, motivo, usuario, fecha, stock_resultante)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (item_id, tipo, cantidad, motivo.strip(), usuario, self.now(), nuevo_stock))
        self.conn.commit()

    def fetch_movements(self, limit=500):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT m.*, i.nombre, i.codigo
            FROM movimientos m
            JOIN insumos i ON i.id = m.insumo_id
            ORDER BY m.fecha DESC, m.id DESC
            LIMIT ?
        """, (limit,))
        return cur.fetchall()

    def fetch_low_stock_items(self):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT * FROM insumos
            WHERE stock_actual <= stock_minimo
            ORDER BY stock_actual ASC, nombre ASC
        """)
        return cur.fetchall()

    def fetch_expiring_items(self, days=30):
        rows = self.fetch_items()
        result = []
        today = date.today()
        for row in rows:
            venc = row["vencimiento"]
            if not venc:
                continue
            try:
                venc_date = datetime.strptime(venc, DATE_FMT).date()
            except ValueError:
                continue
            delta = (venc_date - today).days
            if delta <= days:
                result.append((row, delta))
        result.sort(key=lambda x: x[1])
        return result

    def count_items(self):
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) AS total FROM insumos")
        return cur.fetchone()["total"]

    def total_stock_units(self):
        cur = self.conn.cursor()
        cur.execute("SELECT COALESCE(SUM(stock_actual), 0) AS total FROM insumos")
        return cur.fetchone()["total"]

    def count_movements_today(self):
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) AS total FROM movimientos WHERE date(fecha)=date('now','localtime')")
        return cur.fetchone()["total"]

    def count_expired_items(self):
        cur = self.conn.cursor()
        today = date.today().strftime(DATE_FMT)
        cur.execute("""
            SELECT COUNT(*) AS total
            FROM insumos
            WHERE vencimiento IS NOT NULL AND vencimiento <> '' AND date(vencimiento) < date(?)
        """, (today,))
        return cur.fetchone()["total"]

    def inventory_summary_by_category(self):
        cur = self.conn.cursor()
        cur.execute("""
            SELECT categoria, COUNT(*) AS cantidad_insumos, COALESCE(SUM(stock_actual), 0) AS stock_total
            FROM insumos
            GROUP BY categoria
            ORDER BY categoria
        """)
        return cur.fetchall()

    def export_items_csv(self, filepath, search_text="", category="Todas"):
        rows = self.fetch_items(search_text, category)
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["ID","Código","Nombre","Categoría","Stock actual","Stock mínimo","Unidad","Vencimiento","Ubicación","Proveedor","Lote","Observaciones","Creado en","Actualizado en"])
            for row in rows:
                writer.writerow([row["id"],row["codigo"],row["nombre"],row["categoria"],row["stock_actual"],row["stock_minimo"],row["unidad"],row["vencimiento"],row["ubicacion"],row["proveedor"],row["lote"],row["observaciones"],row["creado_en"],row["actualizado_en"]])

    def export_movements_csv(self, filepath):
        rows = self.fetch_movements(limit=100000)
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["ID","Insumo","Código","Tipo","Cantidad","Motivo","Usuario","Fecha","Stock resultante"])
            for row in rows:
                writer.writerow([row["id"],row["nombre"],row["codigo"],row["tipo"],row["cantidad"],row["motivo"],row["usuario"],row["fecha"],row["stock_resultante"]])

    def export_items_excel(self, filepath, search_text="", category="Todas"):
        if not OPENPYXL_OK:
            raise RuntimeError("Falta instalar openpyxl.")
        rows = self.fetch_items(search_text, category)
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        headers = ["ID","Código","Nombre","Categoría","Stock actual","Stock mínimo","Unidad","Vencimiento","Ubicación","Proveedor","Lote","Observaciones","Creado en","Actualizado en"]
        ws.append(headers)
        for row in rows:
            ws.append([row["id"],row["codigo"],row["nombre"],row["categoria"],row["stock_actual"],row["stock_minimo"],row["unidad"],row["vencimiento"],row["ubicacion"],row["proveedor"],row["lote"],row["observaciones"],row["creado_en"],row["actualizado_en"]])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(horizontal="center")
        widths = [8,14,24,18,12,12,12,16,18,18,14,28,22,22]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        wb.save(filepath)

    def export_movements_excel(self, filepath):
        if not OPENPYXL_OK:
            raise RuntimeError("Falta instalar openpyxl.")
        rows = self.fetch_movements(limit=100000)
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"
        headers = ["ID","Insumo","Código","Tipo","Cantidad","Motivo","Usuario","Fecha","Stock resultante"]
        ws.append(headers)
        for row in rows:
            ws.append([row["id"],row["nombre"],row["codigo"],row["tipo"],row["cantidad"],row["motivo"],row["usuario"],row["fecha"],row["stock_resultante"]])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E4F4E8")
            cell.alignment = Alignment(horizontal="center")
        widths = [8,24,14,12,12,28,18,22,16]
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        wb.save(filepath)

    def export_items_pdf(self, filepath, search_text="", category="Todas"):
        if not REPORTLAB_OK:
            raise RuntimeError("Falta instalar reportlab.")
        rows = self.fetch_items(search_text, category)
        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        data = [["ID","Código","Nombre","Categoría","Stock","Mínimo","Unidad","Vencimiento","Ubicación"]]
        for row in rows:
            data.append([row["id"],row["codigo"] or "",row["nombre"],row["categoria"],row["stock_actual"],row["stock_minimo"],row["unidad"],row["vencimiento"] or "",row["ubicacion"] or ""])
        elements = [Paragraph("Inventario de Insumos Médicos", styles["Title"]), Spacer(1, 8)]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#4F81BD")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.4,colors.grey),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.whitesmoke,colors.HexColor("#EAF2FA")]),
            ("FONTSIZE",(0,0),(-1,-1),8),
        ]))
        elements.append(table)
        doc.build(elements)

    def export_movements_pdf(self, filepath):
        if not REPORTLAB_OK:
            raise RuntimeError("Falta instalar reportlab.")
        rows = self.fetch_movements(limit=300)
        doc = SimpleDocTemplate(filepath, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
        styles = getSampleStyleSheet()
        data = [["ID","Insumo","Código","Tipo","Cantidad","Motivo","Usuario","Fecha","Stock resultante"]]
        for row in rows:
            data.append([row["id"],row["nombre"],row["codigo"] or "",row["tipo"],row["cantidad"],row["motivo"] or "",row["usuario"],row["fecha"],row["stock_resultante"]])
        elements = [Paragraph("Historial de Movimientos", styles["Title"]), Spacer(1, 8)]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#6AA84F")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("GRID",(0,0),(-1,-1),0.4,colors.grey),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.whitesmoke,colors.HexColor("#EEF7EA")]),
            ("FONTSIZE",(0,0),(-1,-1),8),
        ]))
        elements.append(table)
        doc.build(elements)


class InventoryAppV43(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Inventario de Insumos Médicos v4.3")
        self.geometry("1540x920")
        self.minsize(1280, 780)
        self.configure(bg="#F4F7FB")

        self.db = Database()
        self.current_user = None
        self.selected_id = None
        self.fields = {}

        self.main_content = None
        self.content_area = None

        self.search_var = tk.StringVar(value="")
        self.category_var = tk.StringVar(value="Todas")
        self.mov_item_var = tk.StringVar(value="")
        self.mov_type_var = tk.StringVar(value="Entrada")

        self.setup_styles()
        self.create_login_screen()

    def setup_styles(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Treeview", rowheight=28, font=("Segoe UI", 9))
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("Title.TLabel", font=("Segoe UI", 18, "bold"), background="#F4F7FB")
        style.configure("Soft.TLabel", background="#F4F7FB", font=("Segoe UI", 10))
        style.configure("Card.TLabelframe", padding=10)
        style.configure("Card.TLabelframe.Label", font=("Segoe UI", 10, "bold"))

    def clear_window(self):
        for widget in self.winfo_children():
            widget.destroy()

    def create_login_screen(self):
        self.clear_window()
        wrapper = tk.Frame(self, bg="#F4F7FB")
        wrapper.pack(fill="both", expand=True)

        card = tk.Frame(wrapper, bg="white", bd=1, relief="solid")
        card.place(relx=0.5, rely=0.5, anchor="center", width=450, height=350)

        tk.Label(card, text="🏥", bg="white", font=("Segoe UI Emoji", 28)).pack(pady=(18, 6))
        tk.Label(card, text="Inventario de Insumos Médicos v4.3", bg="white", fg="#24476B",
                 font=("Segoe UI", 15, "bold")).pack(pady=(0, 14))

        inner = tk.Frame(card, bg="white")
        inner.pack(fill="both", expand=True, padx=26)

        tk.Label(inner, text="Usuario", bg="white", anchor="w", font=("Segoe UI", 10)).pack(fill="x")
        self.ent_user = ttk.Entry(inner)
        self.ent_user.pack(fill="x", pady=(4, 10))
        self.ent_user.insert(0, DEFAULT_USER)

        tk.Label(inner, text="Contraseña", bg="white", anchor="w", font=("Segoe UI", 10)).pack(fill="x")
        self.ent_pass = ttk.Entry(inner, show="*")
        self.ent_pass.pack(fill="x", pady=(4, 10))
        self.ent_pass.insert(0, DEFAULT_PASS)
        self.ent_pass.bind("<Return>", lambda e: self.try_login())

        tk.Label(inner, text=f"Usuario inicial: {DEFAULT_USER}\nContraseña inicial: {DEFAULT_PASS}",
                 bg="white", fg="#5B6B7A", justify="left", font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 14))
        ttk.Button(inner, text="Ingresar", command=self.try_login).pack(fill="x")

    def try_login(self):
        user = self.db.authenticate_user(self.ent_user.get(), self.ent_pass.get())
        if user is None:
            messagebox.showerror("Acceso denegado", "Usuario o contraseña incorrectos.")
            return
        self.current_user = user
        self.build_main_app()

    def build_main_app(self):
        self.clear_window()

        root = tk.Frame(self, bg="#F4F7FB")
        root.pack(fill="both", expand=True)

        sidebar = tk.Frame(root, bg="#24476B", width=240)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        tk.Label(sidebar, text="🏥", bg="#24476B", fg="white", font=("Segoe UI Emoji", 24)).pack(pady=(20, 6))
        tk.Label(sidebar, text="Inventario Salud", bg="#24476B", fg="white", font=("Segoe UI", 15, "bold")).pack()
        tk.Label(sidebar, text="v4.3", bg="#24476B", fg="#D7E6F5", font=("Segoe UI", 10)).pack(pady=(0, 18))
        tk.Label(sidebar, text=self.current_user["nombre_completo"], bg="#24476B", fg="white", font=("Segoe UI", 10, "bold")).pack(pady=(0, 2))
        tk.Label(sidebar, text=self.current_user["rol"], bg="#24476B", fg="#D7E6F5", font=("Segoe UI", 9)).pack(pady=(0, 18))

        ttk.Button(sidebar, text="Inventario", command=lambda: self.show_section("inventario")).pack(fill="x", padx=18, pady=6)
        ttk.Button(sidebar, text="Movimientos", command=lambda: self.show_section("movimientos")).pack(fill="x", padx=18, pady=6)
        ttk.Button(sidebar, text="Reportes y alertas", command=lambda: self.show_section("reportes")).pack(fill="x", padx=18, pady=6)
        ttk.Button(sidebar, text="Cambiar contraseña", command=self.open_change_password).pack(fill="x", padx=18, pady=6)
        ttk.Button(sidebar, text="Crear respaldo", command=self.create_backup).pack(fill="x", padx=18, pady=6)
        ttk.Button(sidebar, text="Cerrar sesión", command=self.logout).pack(fill="x", padx=18, pady=(24, 6))

        self.main_content = tk.Frame(root, bg="#F4F7FB")
        self.main_content.pack(side="left", fill="both", expand=True)

        self.show_dashboard_shell()
        self.show_section("inventario")

    def show_dashboard_shell(self):
        for w in self.main_content.winfo_children():
            w.destroy()

        header = tk.Frame(self.main_content, bg="#F4F7FB")
        header.pack(fill="x", padx=14, pady=(14, 10))
        ttk.Label(header, text="Inventario de Insumos Médicos v4.3", style="Title.TLabel").pack(side="left")
        ttk.Label(header, text=f"Usuario: {self.current_user['username']}", style="Soft.TLabel").pack(side="right")

        dash = tk.Frame(self.main_content, bg="#F4F7FB")
        dash.pack(fill="x", padx=14, pady=(0, 10))
        self.lbl_total_items = self.make_stat_card(dash, "Productos", "#D9EAF7")
        self.lbl_total_stock = self.make_stat_card(dash, "Unidades", "#E4F4E8")
        self.lbl_low_stock = self.make_stat_card(dash, "Bajo stock", "#FBE5D6")
        self.lbl_expiring = self.make_stat_card(dash, "Por vencer", "#FFF2CC")
        self.lbl_expired = self.make_stat_card(dash, "Vencidos", "#F4CCCC")
        self.lbl_mov_today = self.make_stat_card(dash, "Movimientos hoy", "#EADCF8")

        self.content_area = tk.Frame(self.main_content, bg="#F4F7FB")
        self.content_area.pack(fill="both", expand=True, padx=14, pady=(0, 14))
        self.update_dashboard()

    def make_stat_card(self, parent, title, bg):
        frame = tk.Frame(parent, bg=bg, bd=0)
        frame.pack(side="left", fill="x", expand=True, padx=5)
        tk.Label(frame, text=title, bg=bg, fg="#2F4F6F", font=("Segoe UI", 9, "bold")).pack(pady=(8, 2))
        value = tk.Label(frame, text="0", bg=bg, fg="#1F2937", font=("Segoe UI", 16, "bold"))
        value.pack(pady=(0, 8))
        return value

    def show_section(self, section):
        if self.content_area is None:
            return
        for w in self.content_area.winfo_children():
            w.destroy()
        self.selected_id = None
        if section == "inventario":
            self.create_inventory_section(self.content_area)
        elif section == "movimientos":
            self.create_movements_section(self.content_area)
        elif section == "reportes":
            self.create_reports_section(self.content_area)
        self.refresh_categories()
        self.refresh_table()
        self.refresh_movements_table()
        self.refresh_summary_table()
        self.update_dashboard()

    def create_inventory_section(self, parent):
        self.fields = {}
        parent.columnconfigure(0, weight=0)
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(0, weight=1)

        left = ttk.LabelFrame(parent, text="Formulario de insumo", style="Card.TLabelframe")
        left.grid(row=0, column=0, sticky="ns", padx=(0, 12))

        labels = [
            ("Código", "codigo"),
            ("Nombre", "nombre"),
            ("Categoría", "categoria"),
            ("Stock inicial", "stock_actual"),
            ("Stock mínimo", "stock_minimo"),
            ("Unidad", "unidad"),
            ("Vencimiento (YYYY-MM-DD)", "vencimiento"),
            ("Ubicación", "ubicacion"),
            ("Proveedor", "proveedor"),
            ("Lote", "lote"),
            ("Observaciones", "observaciones"),
        ]

        for i, (text, key) in enumerate(labels):
            ttk.Label(left, text=text).grid(row=i, column=0, sticky="w", pady=4, padx=(0, 8))
            if key == "observaciones":
                widget = tk.Text(left, width=28, height=5, relief="solid", bd=1)
            else:
                widget = ttk.Entry(left, width=30)
            widget.grid(row=i, column=1, sticky="ew", pady=4)
            self.fields[key] = widget

        self.fields["stock_actual"].insert(0, "0")
        self.fields["stock_minimo"].insert(0, "0")

        btn_area = ttk.Frame(left)
        btn_area.grid(row=len(labels), column=0, columnspan=2, sticky="ew", pady=(10, 0))
        btn_area.columnconfigure(0, weight=1)
        btn_area.columnconfigure(1, weight=1)

        ttk.Button(btn_area, text="Nuevo / Limpiar", command=self.clear_form).grid(row=0, column=0, sticky="ew", padx=(0, 4), pady=3)
        ttk.Button(btn_area, text="Guardar insumo", command=self.save_item).grid(row=0, column=1, sticky="ew", padx=(4, 0), pady=3)
        ttk.Button(btn_area, text="Eliminar seleccionado", command=self.delete_selected).grid(row=1, column=0, columnspan=2, sticky="ew", pady=3)
        ttk.Button(btn_area, text="Importar Excel", command=self.import_items_excel).grid(row=2, column=0, columnspan=2, sticky="ew", pady=3)
        ttk.Button(btn_area, text="Descargar plantilla", command=self.download_template_excel).grid(row=3, column=0, columnspan=2, sticky="ew", pady=3)

        ttk.Label(left, text="Exportar inventario").grid(row=len(labels)+1, column=0, sticky="w", pady=(10, 4))
        export_row = ttk.Frame(left)
        export_row.grid(row=len(labels)+2, column=0, columnspan=2, sticky="ew")
        export_row.columnconfigure(0, weight=1)
        export_row.columnconfigure(1, weight=1)
        export_row.columnconfigure(2, weight=1)
        ttk.Button(export_row, text="CSV", command=self.export_items_csv).grid(row=0, column=0, sticky="ew", padx=(0, 4))
        ttk.Button(export_row, text="Excel", command=self.export_items_excel).grid(row=0, column=1, sticky="ew", padx=4)
        ttk.Button(export_row, text="PDF", command=self.export_items_pdf).grid(row=0, column=2, sticky="ew", padx=(4, 0))

        right = ttk.Frame(parent)
        right.grid(row=0, column=1, sticky="nsew")
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        filters = ttk.LabelFrame(right, text="Búsqueda y filtros", style="Card.TLabelframe")
        filters.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        filters.columnconfigure(1, weight=1)

        ttk.Label(filters, text="Buscar").grid(row=0, column=0, sticky="w", padx=(0, 8))
        ent_search = ttk.Entry(filters, textvariable=self.search_var, width=35)
        ent_search.grid(row=0, column=1, sticky="ew", padx=(0, 12))
        ent_search.bind("<KeyRelease>", lambda e: self.refresh_table())

        ttk.Label(filters, text="Categoría").grid(row=0, column=2, sticky="w", padx=(0, 8))
        self.cmb_category = ttk.Combobox(filters, textvariable=self.category_var, state="readonly", width=20)
        self.cmb_category.grid(row=0, column=3, sticky="w", padx=(0, 12))
        self.cmb_category.bind("<<ComboboxSelected>>", lambda e: self.refresh_table())
        ttk.Button(filters, text="Actualizar", command=self.refresh_table).grid(row=0, column=4, sticky="w")

        table_frame = ttk.LabelFrame(right, text="Listado de insumos", style="Card.TLabelframe")
        table_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        columns = ("id","codigo","nombre","categoria","stock_actual","stock_minimo","unidad","vencimiento","ubicacion","proveedor","lote")
        self.tree_items = ttk.Treeview(table_frame, columns=columns, show="headings")
        heads = {"id":"ID","codigo":"Código","nombre":"Nombre","categoria":"Categoría","stock_actual":"Stock","stock_minimo":"Mínimo","unidad":"Unidad","vencimiento":"Vencimiento","ubicacion":"Ubicación","proveedor":"Proveedor","lote":"Lote"}
        widths = {"id":55,"codigo":90,"nombre":180,"categoria":120,"stock_actual":75,"stock_minimo":75,"unidad":90,"vencimiento":110,"ubicacion":110,"proveedor":120,"lote":85}
        for col in columns:
            self.tree_items.heading(col, text=heads[col])
            self.tree_items.column(col, width=widths[col], anchor="center")
        self.tree_items.grid(row=0, column=0, sticky="nsew")
        self.tree_items.bind("<<TreeviewSelect>>", self.on_item_select)
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree_items.yview)
        self.tree_items.configure(yscrollcommand=yscroll.set)
        yscroll.grid(row=0, column=1, sticky="ns")

    def create_movements_section(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)

        top = ttk.LabelFrame(parent, text="Registrar movimiento", style="Card.TLabelframe")
        top.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        ttk.Label(top, text="Insumo").grid(row=0, column=0, sticky="w", padx=(0, 8), pady=4)
        self.cmb_mov_item = ttk.Combobox(top, textvariable=self.mov_item_var, state="readonly", width=45)
        self.cmb_mov_item.grid(row=0, column=1, sticky="w", padx=(0, 12), pady=4)

        ttk.Label(top, text="Tipo").grid(row=0, column=2, sticky="w", padx=(0, 8), pady=4)
        ttk.Combobox(top, textvariable=self.mov_type_var, state="readonly", values=["Entrada", "Salida"], width=15).grid(row=0, column=3, sticky="w", pady=4)

        ttk.Label(top, text="Cantidad").grid(row=1, column=0, sticky="w", padx=(0, 8), pady=4)
        self.ent_mov_qty = ttk.Entry(top, width=18)
        self.ent_mov_qty.grid(row=1, column=1, sticky="w", pady=4)

        ttk.Label(top, text="Motivo").grid(row=1, column=2, sticky="w", padx=(0, 8), pady=4)
        self.ent_mov_reason = ttk.Entry(top, width=28)
        self.ent_mov_reason.grid(row=1, column=3, sticky="w", pady=4)
        ttk.Button(top, text="Registrar movimiento", command=self.register_movement).grid(row=0, column=4, rowspan=2, padx=(12, 0))

        table_frame = ttk.LabelFrame(parent, text="Historial de movimientos", style="Card.TLabelframe")
        table_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        cols = ("id","nombre","codigo","tipo","cantidad","motivo","usuario","fecha","stock_resultante")
        self.tree_mov = ttk.Treeview(table_frame, columns=cols, show="headings")
        heads = {"id":"ID","nombre":"Insumo","codigo":"Código","tipo":"Tipo","cantidad":"Cantidad","motivo":"Motivo","usuario":"Usuario","fecha":"Fecha","stock_resultante":"Stock resultante"}
        widths = {"id":55,"nombre":180,"codigo":95,"tipo":85,"cantidad":80,"motivo":220,"usuario":120,"fecha":150,"stock_resultante":110}
        for col in cols:
            self.tree_mov.heading(col, text=heads[col])
            self.tree_mov.column(col, width=widths[col], anchor="center")
        self.tree_mov.grid(row=0, column=0, sticky="nsew")
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree_mov.yview)
        self.tree_mov.configure(yscrollcommand=yscroll.set)
        yscroll.grid(row=0, column=1, sticky="ns")

        bottom = ttk.Frame(parent)
        bottom.grid(row=2, column=0, sticky="w", pady=(8, 0))
        ttk.Button(bottom, text="CSV", command=self.export_movements_csv).pack(side="left", padx=(0, 5))
        ttk.Button(bottom, text="Excel", command=self.export_movements_excel).pack(side="left", padx=5)
        ttk.Button(bottom, text="PDF", command=self.export_movements_pdf).pack(side="left", padx=5)

    def create_reports_section(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.columnconfigure(1, weight=1)
        parent.rowconfigure(0, weight=1)

        left = ttk.LabelFrame(parent, text="Alertas", style="Card.TLabelframe")
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        ttk.Label(left, text="Insumos con bajo stock").pack(anchor="w", pady=(0, 4))
        self.lst_low = tk.Listbox(left, height=14, relief="solid", bd=1)
        self.lst_low.pack(fill="both", expand=True, pady=(0, 10))

        ttk.Label(left, text="Insumos por vencer o vencidos").pack(anchor="w", pady=(0, 4))
        self.lst_exp = tk.Listbox(left, height=14, relief="solid", bd=1)
        self.lst_exp.pack(fill="both", expand=True)

        right = ttk.LabelFrame(parent, text="Resumen por categoría", style="Card.TLabelframe")
        right.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        right.columnconfigure(0, weight=1)
        right.rowconfigure(0, weight=1)

        cols = ("categoria","cantidad_insumos","stock_total")
        self.tree_summary = ttk.Treeview(right, columns=cols, show="headings", height=18)
        self.tree_summary.heading("categoria", text="Categoría")
        self.tree_summary.heading("cantidad_insumos", text="Cantidad de insumos")
        self.tree_summary.heading("stock_total", text="Stock total")
        self.tree_summary.column("categoria", width=180, anchor="center")
        self.tree_summary.column("cantidad_insumos", width=150, anchor="center")
        self.tree_summary.column("stock_total", width=120, anchor="center")
        self.tree_summary.grid(row=0, column=0, sticky="nsew")

    def get_text_value(self, widget):
        if isinstance(widget, tk.Text):
            return widget.get("1.0", "end").strip()
        return widget.get().strip()

    def set_text_value(self, widget, value):
        value = "" if value is None else str(value)
        if isinstance(widget, tk.Text):
            widget.delete("1.0", "end")
            widget.insert("1.0", value)
        else:
            widget.delete(0, "end")
            widget.insert(0, value)

    def clear_form(self):
        self.selected_id = None
        for key, widget in self.fields.items():
            self.set_text_value(widget, "")
        if "stock_actual" in self.fields:
            self.set_text_value(self.fields["stock_actual"], "0")
        if "stock_minimo" in self.fields:
            self.set_text_value(self.fields["stock_minimo"], "0")

    def validate_item_data(self):
        data = {k: self.get_text_value(w) for k, w in self.fields.items()}
        if not data["nombre"]:
            raise ValueError("El nombre es obligatorio.")
        if not data["categoria"]:
            raise ValueError("La categoría es obligatoria.")
        if not data["unidad"]:
            raise ValueError("La unidad es obligatoria.")
        try:
            data["stock_actual"] = int(data["stock_actual"] or 0)
            data["stock_minimo"] = int(data["stock_minimo"] or 0)
        except ValueError:
            raise ValueError("Stock inicial y stock mínimo deben ser números enteros.")
        if data["stock_actual"] < 0 or data["stock_minimo"] < 0:
            raise ValueError("Los stocks no pueden ser negativos.")
        if data["vencimiento"]:
            try:
                datetime.strptime(data["vencimiento"], DATE_FMT)
            except ValueError:
                raise ValueError("La fecha de vencimiento debe tener formato YYYY-MM-DD.")
        return data

    def save_item(self):
        try:
            data = self.validate_item_data()
            if self.selected_id is None:
                self.db.add_item(data, self.current_user["username"])
                messagebox.showinfo("Éxito", "Insumo registrado correctamente.")
            else:
                self.db.update_item(self.selected_id, data)
                messagebox.showinfo("Éxito", "Insumo actualizado correctamente.")
            self.clear_form()
            self.refresh_categories()
            self.refresh_table()
            self.refresh_movements_table()
            self.refresh_summary_table()
            self.update_dashboard()
        except sqlite3.IntegrityError:
            messagebox.showerror("Error", "El código ya existe. Usa otro código.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def delete_selected(self):
        if self.selected_id is None:
            messagebox.showwarning("Atención", "Selecciona un insumo para eliminar.")
            return
        if not messagebox.askyesno("Confirmar", "¿Seguro que deseas eliminar el insumo seleccionado?"):
            return
        self.db.delete_item(self.selected_id)
        self.clear_form()
        self.refresh_table()
        self.refresh_movements_table()
        self.refresh_summary_table()
        self.update_dashboard()

    def on_item_select(self, event=None):
        if not hasattr(self, "tree_items"):
            return
        selected = self.tree_items.selection()
        if not selected:
            return
        values = self.tree_items.item(selected[0], "values")
        item_id = int(values[0])
        row = self.db.get_item_by_id(item_id)
        if row is None:
            return
        self.selected_id = item_id
        for key in self.fields:
            self.set_text_value(self.fields[key], row[key])

    def refresh_categories(self):
        categories = ["Todas"] + self.db.get_categories()
        if hasattr(self, "cmb_category"):
            self.cmb_category["values"] = categories
            if self.category_var.get() not in categories:
                self.category_var.set("Todas")
        if hasattr(self, "cmb_mov_item"):
            values = [f'{row["id"]} | {row["nombre"]} | Stock: {row["stock_actual"]}' for row in self.db.fetch_items()]
            self.cmb_mov_item["values"] = values
            if values and not self.mov_item_var.get():
                self.mov_item_var.set(values[0])

    def refresh_table(self):
        if not hasattr(self, "tree_items"):
            return
        for item in self.tree_items.get_children():
            self.tree_items.delete(item)
        for row in self.db.fetch_items(self.search_var.get(), self.category_var.get()):
            self.tree_items.insert("", "end", values=(row["id"],row["codigo"],row["nombre"],row["categoria"],row["stock_actual"],row["stock_minimo"],row["unidad"],row["vencimiento"],row["ubicacion"],row["proveedor"],row["lote"]))

    def refresh_movements_table(self):
        if not hasattr(self, "tree_mov"):
            return
        for item in self.tree_mov.get_children():
            self.tree_mov.delete(item)
        for row in self.db.fetch_movements():
            self.tree_mov.insert("", "end", values=(row["id"],row["nombre"],row["codigo"],row["tipo"],row["cantidad"],row["motivo"],row["usuario"],row["fecha"],row["stock_resultante"]))

    def refresh_summary_table(self):
        if hasattr(self, "tree_summary"):
            for item in self.tree_summary.get_children():
                self.tree_summary.delete(item)
            for row in self.db.inventory_summary_by_category():
                self.tree_summary.insert("", "end", values=(row["categoria"], row["cantidad_insumos"], row["stock_total"]))
        if hasattr(self, "lst_low"):
            self.lst_low.delete(0, "end")
            for row in self.db.fetch_low_stock_items():
                self.lst_low.insert("end", f'{row["nombre"]} | stock {row["stock_actual"]} | mínimo {row["stock_minimo"]}')
        if hasattr(self, "lst_exp"):
            self.lst_exp.delete(0, "end")
            for row, delta in self.db.fetch_expiring_items(days=30):
                estado = "vencido" if delta < 0 else f"vence en {delta} días"
                self.lst_exp.insert("end", f'{row["nombre"]} | {row["vencimiento"]} | {estado}')

    def update_dashboard(self):
        if hasattr(self, "lbl_total_items"):
            self.lbl_total_items.config(text=str(self.db.count_items()))
            self.lbl_total_stock.config(text=str(self.db.total_stock_units()))
            self.lbl_low_stock.config(text=str(len(self.db.fetch_low_stock_items())))
            self.lbl_expiring.config(text=str(len([x for x in self.db.fetch_expiring_items(days=30) if x[1] >= 0])))
            self.lbl_expired.config(text=str(self.db.count_expired_items()))
            self.lbl_mov_today.config(text=str(self.db.count_movements_today()))

    def register_movement(self):
        text = self.mov_item_var.get().strip()
        if not text:
            messagebox.showwarning("Atención", "Selecciona un insumo.")
            return
        try:
            item_id = int(text.split("|")[0].strip())
        except Exception:
            messagebox.showerror("Error", "No se pudo identificar el insumo.")
            return
        try:
            qty = int(self.ent_mov_qty.get().strip())
        except ValueError:
            messagebox.showerror("Error", "La cantidad debe ser un número entero.")
            return
        reason = self.ent_mov_reason.get().strip() or "Sin motivo especificado"
        try:
            self.db.add_movement(item_id, self.mov_type_var.get(), qty, reason, self.current_user["username"])
            messagebox.showinfo("Éxito", "Movimiento registrado correctamente.")
            self.ent_mov_qty.delete(0, "end")
            self.ent_mov_reason.delete(0, "end")
            self.refresh_categories()
            self.refresh_table()
            self.refresh_movements_table()
            self.refresh_summary_table()
            self.update_dashboard()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def save_dialog(self, title, ext):
        return filedialog.asksaveasfilename(title=title, defaultextension=ext, filetypes=[(ext.upper().replace(".", ""), f"*{ext}")])

    def export_items_csv(self):
        path = self.save_dialog("Guardar inventario CSV", ".csv")
        if path:
            try:
                self.db.export_items_csv(path, self.search_var.get(), self.category_var.get())
                messagebox.showinfo("Éxito", "Inventario exportado en CSV.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def export_items_excel(self):
        path = self.save_dialog("Guardar inventario Excel", ".xlsx")
        if path:
            try:
                self.db.export_items_excel(path, self.search_var.get(), self.category_var.get())
                messagebox.showinfo("Éxito", "Inventario exportado en Excel.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def export_items_pdf(self):
        path = self.save_dialog("Guardar inventario PDF", ".pdf")
        if path:
            try:
                self.db.export_items_pdf(path, self.search_var.get(), self.category_var.get())
                messagebox.showinfo("Éxito", "Inventario exportado en PDF.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def export_movements_csv(self):
        path = self.save_dialog("Guardar movimientos CSV", ".csv")
        if path:
            try:
                self.db.export_movements_csv(path)
                messagebox.showinfo("Éxito", "Movimientos exportados en CSV.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def export_movements_excel(self):
        path = self.save_dialog("Guardar movimientos Excel", ".xlsx")
        if path:
            try:
                self.db.export_movements_excel(path)
                messagebox.showinfo("Éxito", "Movimientos exportados en Excel.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def export_movements_pdf(self):
        path = self.save_dialog("Guardar movimientos PDF", ".pdf")
        if path:
            try:
                self.db.export_movements_pdf(path)
                messagebox.showinfo("Éxito", "Movimientos exportados en PDF.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def download_template_excel(self):
        if not OPENPYXL_OK:
            messagebox.showerror("Error", "Falta instalar openpyxl.")
            return
        path = self.save_dialog("Guardar plantilla Excel", ".xlsx")
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Plantilla Medicamentos"
            headers = ["codigo","nombre","categoria","stock_actual","stock_minimo","unidad","vencimiento","ubicacion","proveedor","lote","observaciones"]
            ws.append(headers)
            ws.append(["MED-001","Paracetamol 500 mg","Analgésico",120,20,"caja","2027-12-31","Bodega A","Proveedor Salud","L001","Uso adulto"])
            ws.append(["MED-002","Ibuprofeno 400 mg","Antiinflamatorio",80,15,"caja","2027-10-15","Bodega A","Proveedor Salud","L002","Tomar con comida"])
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(horizontal="center")
            widths = [14,28,20,14,14,14,16,18,22,14,28]
            for i, width in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
            wb.save(path)
            messagebox.showinfo("Éxito", "Plantilla Excel guardada correctamente.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def import_items_excel(self):
        if not OPENPYXL_OK:
            messagebox.showerror("Error", "Falta instalar openpyxl.")
            return
        path = filedialog.askopenfilename(title="Seleccionar archivo Excel", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            wb = load_workbook(path)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
            required = ["codigo","nombre","categoria","stock_actual","stock_minimo","unidad","vencimiento","ubicacion","proveedor","lote","observaciones"]
            for col in required:
                if col not in headers:
                    raise ValueError(f"Falta la columna requerida: {col}")
            idx = {h: headers.index(h) for h in required}
            imported = 0
            updated = 0

            for row_num in range(2, ws.max_row + 1):
                values = {}
                empty = True
                for col in required:
                    cell_val = ws.cell(row=row_num, column=idx[col] + 1).value
                    if cell_val is not None and str(cell_val).strip() != "":
                        empty = False
                    values[col] = "" if cell_val is None else str(cell_val).strip()
                if empty:
                    continue

                try:
                    values["stock_actual"] = int(float(values["stock_actual"] or 0))
                    values["stock_minimo"] = int(float(values["stock_minimo"] or 0))
                except ValueError:
                    raise ValueError(f"Fila {row_num}: stock_actual y stock_minimo deben ser numéricos.")

                data = {
                    "codigo": values["codigo"],
                    "nombre": values["nombre"],
                    "categoria": values["categoria"],
                    "stock_actual": values["stock_actual"],
                    "stock_minimo": values["stock_minimo"],
                    "unidad": values["unidad"],
                    "vencimiento": values["vencimiento"],
                    "ubicacion": values["ubicacion"],
                    "proveedor": values["proveedor"],
                    "lote": values["lote"],
                    "observaciones": values["observaciones"],
                }

                existing = self.db.get_item_by_code(values["codigo"])
                if existing is None:
                    self.db.add_item(data, self.current_user["username"])
                    imported += 1
                else:
                    old_stock = int(existing["stock_actual"])
                    self.db.update_item(existing["id"], data)
                    self.db.set_stock_direct(existing["id"], old_stock)
                    new_stock = int(values["stock_actual"])
                    if new_stock != old_stock:
                        diff = new_stock - old_stock
                        tipo = "Entrada" if diff > 0 else "Salida"
                        self.db.add_movement(existing["id"], tipo, abs(diff), "Ajuste por importación Excel", self.current_user["username"])
                    updated += 1

            self.refresh_categories()
            self.refresh_table()
            self.refresh_movements_table()
            self.refresh_summary_table()
            self.update_dashboard()
            messagebox.showinfo("Éxito", f"Importación completada.\nNuevos: {imported}\nActualizados: {updated}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def open_change_password(self):
        win = tk.Toplevel(self)
        win.title("Cambiar contraseña")
        win.geometry("380x240")
        win.resizable(False, False)
        win.grab_set()

        frame = ttk.Frame(win, padding=18)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text="Contraseña actual").pack(anchor="w")
        ent_current = ttk.Entry(frame, show="*")
        ent_current.pack(fill="x", pady=(4, 10))

        ttk.Label(frame, text="Nueva contraseña").pack(anchor="w")
        ent_new = ttk.Entry(frame, show="*")
        ent_new.pack(fill="x", pady=(4, 10))

        ttk.Label(frame, text="Confirmar nueva contraseña").pack(anchor="w")
        ent_confirm = ttk.Entry(frame, show="*")
        ent_confirm.pack(fill="x", pady=(4, 14))

        def save_password():
            current = ent_current.get().strip()
            new = ent_new.get().strip()
            confirm = ent_confirm.get().strip()
            if new != confirm:
                messagebox.showerror("Error", "La confirmación no coincide.", parent=win)
                return
            try:
                self.db.change_password(self.current_user["username"], current, new)
                messagebox.showinfo("Éxito", "Contraseña actualizada correctamente.", parent=win)
                win.destroy()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=win)

        ttk.Button(frame, text="Guardar cambios", command=save_password).pack(fill="x")

    def create_backup(self):
        try:
            path = self.db.create_backup()
            messagebox.showinfo("Éxito", f"Respaldo creado correctamente.\n\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def logout(self):
        if messagebox.askyesno("Cerrar sesión", "¿Deseas cerrar la sesión actual?"):
            self.current_user = None
            self.selected_id = None
            self.fields = {}
            self.search_var.set("")
            self.category_var.set("Todas")
            self.mov_item_var.set("")
            self.mov_type_var.set("Entrada")
            self.create_login_screen()


if __name__ == "__main__":
    app = InventoryAppV43()
    app.mainloop()
